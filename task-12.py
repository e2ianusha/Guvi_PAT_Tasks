import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class OrangeHRMLoginPage:
    URL = "https://opensource-demo.orangehrmlive.com/"
    USERNAME_FIELD = (By.ID, "txtUsername")
    PASSWORD_FIELD = (By.ID, "txtPassword")
    LOGIN_BUTTON = (By.ID, "btnLogin")
    INVALID_LOGIN_MESSAGE = (By.ID, "spanMessage")

    def __init__(self, driver):
        self.driver = driver

    def open(self):
        self.driver.get(self.URL)

    def login(self, username, password):
        username_field = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.USERNAME_FIELD))
        username_field.clear()
        username_field.send_keys(username)

        password_field = self.driver.find_element(*self.PASSWORD_FIELD)
        password_field.clear()
        password_field.send_keys(password)

        login_button = self.driver.find_element(*self.LOGIN_BUTTON)
        login_button.click()

    def get_invalid_login_message(self):
        return self.driver.find_element(*self.INVALID_LOGIN_MESSAGE).text


@pytest.fixture()
def setup():
    driver = webdriver.Chrome()  # Assuming Chrome WebDriver is used
    yield driver
    driver.quit()


def read_excel_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            "Test ID": row[0],
            "Username": row[1],
            "Password": row[2],
            "Expected Result": row[6]
        })
    return data


def write_to_excel(file_path, test_id, result):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        if row[0].value == test_id:
            sheet.cell(row=row[0].row, column=5).value = result
            workbook.save(file_path)
            return True
    return False


@pytest.mark.parametrize("data", read_excel_data("login_data.xlsx"))
def test_login(setup, data):
    driver = setup
    login_page = OrangeHRMLoginPage(driver)
    login_page.open()

    username = data["Username"]
    password = data["Password"]
    expected_result = data["Expected Result"]
    test_id = data["Test ID"]

    login_page.login(username, password)

    # Check if login was successful
    try:
        WebDriverWait(driver, 10).until(
            EC.url_to_be("https://opensource-demo.orangehrmlive.com/index.php/dashboard"))
        actual_result = "Pass"
    except:
        invalid_message = login_page.get_invalid_login_message()
        if "Invalid credentials" in invalid_message:
            actual_result = "Fail - Invalid credentials"
        else:
            actual_result = "Fail - Unknown error"

    # Write test result to Excel
    write_to_excel("login_data.xlsx", test_id, actual_result)
